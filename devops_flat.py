from openpyxl import Workbook
from test import ProjectTasks, ProjectDict
from download_project import date
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from operator import itemgetter

url = 'https://spanorsi.lancloud.ru/pwa/_api/ProjectData/Задачи?$format=json'
select = '&$select=НазваниеЗадачи,Статусзадачи,ДатаНачалаЗадачи,ДатаОкончанияЗадачи,ЗадачаЯвляетсяСуммарной,УровеньСтруктурыЗадачи,ИндексЗадачи,ПроцентЗавершенияЗадачи,ИдентификаторРодительскойЗадачи,ИдентификаторЗадачи'
id_project = '7076c966-340b-ec11-995f-14dae9ca5f43'
filter_project = """&$filter=ЗадачаЯвляетсяАктивной eq true and ИДПроекта eq guid'"""


def CreateWorkbook(path):
    for project in ProjectDict():
        print(project['ИмяПроекта'])
        id_project = project['ИДПроекта']
        name_project = project['ИмяПроекта']
        today = datetime.today().date()
        wb = Workbook()
        ws = wb.active
        greenFill = PatternFill(start_color='5AF2A1', end_color='5AF2A1', fill_type='solid')
        greenFill2 = PatternFill(start_color='77B89D', end_color='77B89D', fill_type='solid')
        redFill = PatternFill(start_color='F08080', end_color='F08080', fill_type='solid')
        yellowFill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
        blueFill = PatternFill(start_color='95D2F5', end_color='95D2F5', fill_type='solid')
        grayFill = PatternFill(start_color='CED7DE', end_color='CED7DE', fill_type='solid')
        tasks = ProjectTasks(url, select, filter_project + id_project + "'")
        task = []
        # ws.cell(row=1, column=1, value='Индекс задачи')
        # ws.cell(row=1, column=2, value='Уровень структуры задачи')
        ws.cell(row=1, column=1, value='Наименование задачи')
        ws.cell(row=1, column=2, value='Начало')
        ws.cell(row=1, column=3, value='Окончание')
        ws.cell(row=1, column=4, value='Статус задачи')
        ws.cell(row=1, column=5, value='% Факт')
        ws.column_dimensions['A'].width = 70
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 7
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['G'].width = 14
        ws['A1'].fill, ws['A1'].font, ws['A1'].alignment = grayFill, Font(bold=True), Alignment(horizontal='center')
        ws['B1'].fill, ws['B1'].font, ws['B1'].alignment = grayFill, Font(bold=True), Alignment(horizontal='center')
        ws['C1'].fill, ws['C1'].font, ws['C1'].alignment = grayFill, Font(bold=True), Alignment(horizontal='center')
        ws['D1'].fill, ws['D1'].font, ws['D1'].alignment = grayFill, Font(bold=True), Alignment(horizontal='center')
        ws['E1'].fill, ws['E1'].font, ws['E1'].alignment = grayFill, Font(bold=True), Alignment(horizontal='center')
        ws['A3'].fill, ws['A3'].font, ws['A3'].alignment = greenFill2, Font(bold=True), Alignment(horizontal='center',
                                                                                                  vertical='center')
        ws['B3'].fill, ws['B3'].font, ws['B3'].alignment = greenFill2, Font(bold=True), Alignment(horizontal='center',
                                                                                                  vertical='center')
        ws['C3'].fill, ws['C3'].font, ws['C3'].alignment = greenFill2, Font(bold=True), Alignment(horizontal='center',
                                                                                                  vertical='center')
        ws['D3'].fill, ws['D3'].font, ws['D3'].alignment = greenFill2, Font(bold=True), Alignment(horizontal='center',
                                                                                                  vertical='center')
        ws['E3'].fill, ws['E3'].font, ws['E3'].alignment = greenFill2, Font(bold=True), Alignment(horizontal='center',
                                                                                                  vertical='center')

        for i in tasks:
            for j in i:
                task.append(j)

        tasktodeletesum = []
        tasktodeletesum3 = []

        for i in task:
            if i['ЗадачаЯвляетсяСуммарной'] is True and i['УровеньСтруктурыЗадачи'] > 3 and i[
                'ПроцентЗавершенияЗадачи'] == 100:
                tasktodeletesum.append(i['ИдентификаторЗадачи'])
            if i['ЗадачаЯвляетсяСуммарной'] is True and i['УровеньСтруктурыЗадачи'] == 3 and i[
                'ПроцентЗавершенияЗадачи'] == 100:
                tasktodeletesum3.append(i['ИдентификаторЗадачи'])

        tasktodelete = []

        for i in task:
            if i['ИдентификаторРодительскойЗадачи'] in (tasktodeletesum + tasktodeletesum3):
                tasktodelete.append(i['ИдентификаторЗадачи'])

        resultdeletetusk = tasktodeletesum + tasktodelete

        task = [i for i in task if i['ИдентификаторЗадачи'] not in resultdeletetusk]

        task = sorted(task, key=itemgetter('ИндексЗадачи'))

        for i, j in enumerate(task, 1):
            # taskIndex = ws.cell(row=i + 1, column=1, value=j['ИндексЗадачи'])
            # StructurLevel = ws.cell(row=i + 1, column=2, value=j['УровеньСтруктурыЗадачи'])
            NameTask = ws.cell(row=i + 1, column=1, value=j['НазваниеЗадачи'])
            NameTask.alignment = Alignment(wrapText=True)
            StartDate = ws.cell(row=i + 1, column=2, value=date(j['ДатаНачалаЗадачи']))
            StartDate.alignment = Alignment(horizontal='center')
            EndDate = ws.cell(row=i + 1, column=3, value=date(j['ДатаОкончанияЗадачи']))
            EndDate.alignment = Alignment(horizontal='center')
            Status = ws.cell(row=i + 1, column=4, value=j['Статусзадачи'])
            Status.alignment = Alignment(horizontal='center')
            Percent = ws.cell(row=i + 1, column=5, value=j['ПроцентЗавершенияЗадачи'])
            Percent.alignment = Alignment(horizontal='center')
            # ws.cell(row=i + 1, column=8, value=j['ЗадачаЯвляетсяСуммарной'])
            # ws.cell(row=i + 1, column=9, value=j['ИдентификаторРодительскойЗадачи'])
            # ws.cell(row=i + 1, column=10, value=j['ИдентификаторЗадачи'])
            if j['УровеньСтруктурыЗадачи'] == 2:
                # taskIndex.fill = blueFill
                # StructurLevel.fill = blueFill
                NameTask.fill, NameTask.alignment = blueFill, Alignment(indent=2)
                StartDate.fill = blueFill
                EndDate.fill = blueFill
                Status.fill = blueFill
                Percent.fill = blueFill
            if j['УровеньСтруктурыЗадачи'] > 2:
                if j['ЗадачаЯвляетсяСуммарной'] is True:
                    # taskIndex.font = Font(bold=True)
                    # StructurLevel.font = Font(bold=True)
                    NameTask.font, NameTask.alignment = Font(bold=True), Alignment(wrapText=True, indent=int(j[
                                                                                                                 'УровеньСтруктурыЗадачи']
                                                                                                             + 1))
                    StartDate.font = Font(bold=True)
                    EndDate.font = Font(bold=True)
                    Status.font = Font(bold=True)
                    Percent.font = Font(bold=True)
                else:
                    NameTask.alignment = Alignment(indent=int(j['УровеньСтруктурыЗадачи']) + 1)

            date_string = j['ДатаОкончанияЗадачи']
            date_task = datetime.strptime(date_string, "%Y-%m-%dT%H:%M:%S")
            date_task = date_task.date()
            if j['ЗадачаЯвляетсяСуммарной'] is False:

                if date_task <= today and j['ПроцентЗавершенияЗадачи'] != 100:
                    EndDate.fill = redFill

                if date_task > today and 0 < j['ПроцентЗавершенияЗадачи'] < 100:
                    Percent.fill = greenFill

                if j['ПроцентЗавершенияЗадачи'] == 0 and date_task < today:
                    StartDate.fill = yellowFill
        ws.delete_rows(2)
        wb.save(f'{datetime.now().date()}-{path}-{name_project}.xlsx')


CreateWorkbook('Actualisation.xlsx')

# print(datetime.today().date())
# print(type(datetime.today().date()))
